from datetime import datetime, date, timedelta
from typing import List, Dict, Any, Optional
from uuid import UUID
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select
from sqlalchemy.orm import selectinload
from sqlalchemy import func, and_

from app.models.business import Customer, Supplier, Branch, Company
from app.models.product import Product, ProductCategory
from app.models.purchase import PurchaseOrder, PurchaseEntry
from app.models.inventory import CurrentStock, StockTransaction
from app.models.sales import SalesOrder, Invoice, InvoiceItem
from app.models.finance import Payment
from app.schemas.transaction import (
    DashboardResponse, KPICardsOut, SalesByCategoryOut,
    MonthlySalesTrendOut, TopProductSalesOut, RecentTransactionOut,
    LedgerEntry, CustomerLedgerResponse, SupplierLedgerResponse
)


class ReportService:
    @staticmethod
    async def get_dashboard_metrics(db: AsyncSession, branch_id: Optional[UUID] = None) -> DashboardResponse:
        """Aggregate data to populate the primary KPI dashboard charts and tables."""
        today = date.today()
        start_of_today = datetime(today.year, today.month, today.day)
        start_of_month = datetime(today.year, today.month, 1)

        # ---------------------------------------------
        # 1. KPI CARDS SUMS
        # ---------------------------------------------
        # Today's Sales
        stmt_today = select(func.sum(Invoice.total_amount)).filter(Invoice.date >= start_of_today)
        if branch_id:
            stmt_today = stmt_today.filter(Invoice.branch_id == branch_id)
        q_today = await db.execute(stmt_today)
        today_sales = q_today.scalar() or 0.0

        # Monthly Sales
        stmt_month = select(func.sum(Invoice.total_amount)).filter(Invoice.date >= start_of_month)
        if branch_id:
            stmt_month = stmt_month.filter(Invoice.branch_id == branch_id)
        q_month = await db.execute(stmt_month)
        monthly_sales = q_month.scalar() or 0.0

        # Outstanding Payments
        stmt_out = select(func.sum(Invoice.total_amount)).filter(Invoice.status != "Paid")
        if branch_id:
            stmt_out = stmt_out.filter(Invoice.branch_id == branch_id)
        q_out = await db.execute(stmt_out)
        outstanding = q_out.scalar() or 0.0

        # Low Stock count
        # Select products where min_stock_level > current_stock
        stmt_low = select(func.count(Product.id)).outerjoin(
            CurrentStock, and_(CurrentStock.product_id == Product.id, CurrentStock.branch_id == branch_id) if branch_id else CurrentStock.product_id == Product.id
        ).filter(Product.min_stock_level > func.coalesce(CurrentStock.qty, 0.0))
        q_low = await db.execute(stmt_low)
        low_stock = q_low.scalar() or 0

        kpis = KPICardsOut(
            today_sales=today_sales,
            monthly_sales=monthly_sales,
            outstanding_payments=outstanding,
            low_stock_count=low_stock
        )

        # ---------------------------------------------
        # 2. SALES BY CATEGORY PIE CHART
        # ---------------------------------------------
        stmt_cat = (
            select(ProductCategory.name, func.sum(InvoiceItem.amount))
            .join(Product, Product.category_id == ProductCategory.id)
            .join(InvoiceItem, InvoiceItem.product_id == Product.id)
            .join(Invoice, Invoice.id == InvoiceItem.invoice_id)
        )
        if branch_id:
            stmt_cat = stmt_cat.filter(Invoice.branch_id == branch_id)
        stmt_cat = stmt_cat.group_by(ProductCategory.name)
        q_cat = await db.execute(stmt_cat)
        
        sales_by_category = [
            SalesByCategoryOut(category_name=row[0], total_sales=row[1] or 0.0)
            for row in q_cat.all()
        ]

        # ---------------------------------------------
        # 3. MONTHLY SALES TREND (LAST 12 MONTHS)
        # ---------------------------------------------
        # In a real system, we'd group by month using dates. Let's return realistic aggregated metrics.
        monthly_sales_trend = []
        months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
        
        # Pull invoice sums grouped by month
        for idx, m_name in enumerate(months):
            stmt_trend = select(func.sum(Invoice.total_amount)).filter(func.extract('month', Invoice.date) == (idx + 1))
            if branch_id:
                stmt_trend = stmt_trend.filter(Invoice.branch_id == branch_id)
            q_trend = await db.execute(stmt_trend)
            monthly_sales_trend.append(
                MonthlySalesTrendOut(month=m_name, sales=q_trend.scalar() or 0.0)
            )

        # ---------------------------------------------
        # 4. TOP 10 PRODUCTS BY SALES VOLUME
        # ---------------------------------------------
        stmt_top = (
            select(Product.name, Product.sku, func.sum(InvoiceItem.qty), func.sum(InvoiceItem.amount))
            .join(InvoiceItem, InvoiceItem.product_id == Product.id)
            .join(Invoice, Invoice.id == InvoiceItem.invoice_id)
        )
        if branch_id:
            stmt_top = stmt_top.filter(Invoice.branch_id == branch_id)
        stmt_top = stmt_top.group_by(Product.name, Product.sku).order_by(func.sum(InvoiceItem.amount).desc()).limit(10)
        q_top = await db.execute(stmt_top)
        
        top_products = [
            TopProductSalesOut(product_name=row[0], sku=row[1], qty_sold=row[2] or 0.0, total_revenue=row[3] or 0.0)
            for row in q_top.all()
        ]

        # ---------------------------------------------
        # 5. RECENT TRANSACTIONS LOG
        # ---------------------------------------------
        # Fetch last 5 Sales Orders
        stmt_so = select(SalesOrder).options(selectinload(SalesOrder.customer)).order_by(SalesOrder.date.desc()).limit(5)
        if branch_id:
            stmt_so = stmt_so.filter(SalesOrder.branch_id == branch_id)
        q_so = await db.execute(stmt_so)
        
        # Fetch last 5 Purchase Orders
        stmt_po = select(PurchaseOrder).options(selectinload(PurchaseOrder.supplier)).order_by(PurchaseOrder.date.desc()).limit(5)
        if branch_id:
            stmt_po = stmt_po.filter(PurchaseOrder.branch_id == branch_id)
        q_po = await db.execute(stmt_po)

        recent_txs = []
        for so in q_so.scalars().all():
            recent_txs.append(
                RecentTransactionOut(
                    id=so.id,
                    tx_type="Sales Order",
                    reference_no=f"SO-{so.id.hex[:6].upper()}",
                    party_name=so.customer.name,
                    date=so.date,
                    amount=so.grand_total,
                    status=so.status
                )
            )
        
        for po in q_po.scalars().all():
            recent_txs.append(
                RecentTransactionOut(
                    id=po.id,
                    tx_type="Purchase Order",
                    reference_no=f"PO-{po.id.hex[:6].upper()}",
                    party_name=po.supplier.name,
                    date=po.date,
                    amount=po.grand_total,
                    status=po.status
                )
            )
        
        # Sort combined list by date desc
        recent_txs.sort(key=lambda x: x.date, reverse=True)
        recent_transactions = recent_txs[:10]

        return DashboardResponse(
            kpis=kpis,
            sales_by_category=sales_by_category,
            monthly_sales_trend=monthly_sales_trend,
            top_products=top_products,
            recent_transactions=recent_transactions
        )

    @staticmethod
    async def get_customer_ledger(
        db: AsyncSession,
        customer_id: UUID,
        start_date: Optional[str] = None,
        end_date: Optional[str] = None
    ) -> CustomerLedgerResponse:
        # 1. Parse start and end datetimes
        start_dt = None
        end_dt = None
        if start_date:
            try:
                if len(start_date) == 7:
                    start_dt = datetime.strptime(start_date, "%Y-%m")
                else:
                    start_dt = datetime.strptime(start_date, "%Y-%m-%d")
            except Exception:
                pass
        if end_date:
            try:
                if len(end_date) == 7:
                    year, month = map(int, end_date.split("-"))
                    if month == 12:
                        end_dt = datetime(year + 1, 1, 1) - timedelta(seconds=1)
                    else:
                        end_dt = datetime(year, month + 1, 1) - timedelta(seconds=1)
                else:
                    end_dt = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1) - timedelta(seconds=1)
            except Exception:
                pass

        # 2. Fetch Invoices
        stmt_inv = (
            select(Invoice)
            .join(SalesOrder, Invoice.sales_order_id == SalesOrder.id)
            .filter(SalesOrder.customer_id == customer_id, Invoice.status != "Cancelled")
        )
        if start_dt:
            stmt_inv = stmt_inv.filter(Invoice.date >= start_dt)
        if end_dt:
            stmt_inv = stmt_inv.filter(Invoice.date <= end_dt)
        q_inv = await db.execute(stmt_inv.order_by(Invoice.date.asc()))
        invoices = q_inv.scalars().all()

        # 3. Fetch Payments
        stmt_pay = select(Payment).filter(Payment.customer_id == customer_id)
        if start_dt:
            stmt_pay = stmt_pay.filter(Payment.payment_date >= start_dt)
        if end_dt:
            stmt_pay = stmt_pay.filter(Payment.payment_date <= end_dt)
        q_pay = await db.execute(stmt_pay.order_by(Payment.payment_date.asc()))
        payments = q_pay.scalars().all()

        # 4. Merge entries chronologically
        entries = []
        for inv in invoices:
            entries.append({
                "date": inv.date,
                "tx_type": "Invoice",
                "reference_no": inv.invoice_number,
                "debit": inv.total_amount,
                "credit": 0.0
            })
        for pay in payments:
            entries.append({
                "date": pay.payment_date,
                "tx_type": "Payment",
                "reference_no": pay.reference_number or "N/A",
                "debit": 0.0,
                "credit": pay.amount_paid
            })

        # Sort entries by date
        entries.sort(key=lambda x: x["date"])

        # Calculate running balance and totals
        total_billed = 0.0
        total_paid = 0.0
        running_balance = 0.0
        ledger_entries = []

        for e in entries:
            debit = e["debit"]
            credit = e["credit"]
            total_billed += debit
            total_paid += credit
            running_balance += (debit - credit)
            ledger_entries.append(
                LedgerEntry(
                    date=e["date"],
                    tx_type=e["tx_type"],
                    reference_no=e["reference_no"],
                    debit=debit,
                    credit=credit,
                    running_balance=running_balance
                )
            )

        return CustomerLedgerResponse(
            total_billed=total_billed,
            total_paid=total_paid,
            balance=running_balance,
            transactions=ledger_entries
        )

    @staticmethod
    async def get_supplier_ledger(
        db: AsyncSession,
        supplier_id: UUID,
        start_date: Optional[str] = None,
        end_date: Optional[str] = None
    ) -> SupplierLedgerResponse:
        # 1. Parse start and end datetimes
        start_dt = None
        end_dt = None
        if start_date:
            try:
                if len(start_date) == 7:
                    start_dt = datetime.strptime(start_date, "%Y-%m")
                else:
                    start_dt = datetime.strptime(start_date, "%Y-%m-%d")
            except Exception:
                pass
        if end_date:
            try:
                if len(end_date) == 7:
                    year, month = map(int, end_date.split("-"))
                    if month == 12:
                        end_dt = datetime(year + 1, 1, 1) - timedelta(seconds=1)
                    else:
                        end_dt = datetime(year, month + 1, 1) - timedelta(seconds=1)
                else:
                    end_dt = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1) - timedelta(seconds=1)
            except Exception:
                pass

        # 2. Fetch Bills (PurchaseEntry)
        stmt_bills = select(PurchaseEntry).filter(PurchaseEntry.supplier_id == supplier_id, PurchaseEntry.status != "Cancelled")
        if start_dt:
            stmt_bills = stmt_bills.filter(PurchaseEntry.billing_date >= start_dt)
        if end_dt:
            stmt_bills = stmt_bills.filter(PurchaseEntry.billing_date <= end_dt)
        q_bills = await db.execute(stmt_bills.order_by(PurchaseEntry.billing_date.asc()))
        bills = q_bills.scalars().all()

        # 3. Construct ledger entries
        entries = []
        for bill in bills:
            paid_amt = bill.total_amount if bill.status == "Paid" else 0.0
            entries.append({
                "date": bill.billing_date,
                "tx_type": "Purchase Bill",
                "reference_no": bill.invoice_number,
                "debit": bill.total_amount,
                "credit": paid_amt
            })

        # Sort entries by date
        entries.sort(key=lambda x: x["date"])

        total_purchased = 0.0
        total_paid = 0.0
        running_balance = 0.0
        ledger_entries = []

        for e in entries:
            debit = e["debit"]
            credit = e["credit"]
            total_purchased += debit
            total_paid += credit
            running_balance += (debit - credit)
            ledger_entries.append(
                LedgerEntry(
                    date=e["date"],
                    tx_type=e["tx_type"],
                    reference_no=e["reference_no"],
                    debit=debit,
                    credit=credit,
                    running_balance=running_balance
                )
            )

        return SupplierLedgerResponse(
            total_purchased=total_purchased,
            total_paid=total_paid,
            balance=running_balance,
            transactions=ledger_entries
        )

    @staticmethod
    async def get_sales_summary_data(
        db: AsyncSession,
        start_date: Optional[str] = None,
        end_date: Optional[str] = None,
        branch_id: Optional[UUID] = None
    ) -> List[Dict[str, Any]]:
        """Fetch and format sales summary item level data for Excel / UI report."""
        stmt = (
            select(Invoice)
            .options(
                selectinload(Invoice.items).selectinload(InvoiceItem.product),
                selectinload(Invoice.sales_order).selectinload(SalesOrder.customer),
                selectinload(Invoice.payments),
                selectinload(Invoice.delivery_challan)
            )
            .filter(Invoice.status != "Cancelled")
        )
        
        if branch_id:
            stmt = stmt.filter(Invoice.branch_id == branch_id)
            
        if start_date:
            try:
                start_dt = datetime.strptime(start_date, "%Y-%m-%d")
                stmt = stmt.filter(Invoice.date >= start_dt)
            except ValueError:
                pass
                
        if end_date:
            try:
                end_dt = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1) - timedelta(seconds=1)
                stmt = stmt.filter(Invoice.date <= end_dt)
            except ValueError:
                pass
                
        query = await db.execute(stmt.order_by(Invoice.invoice_number.asc()))
        invoices = list(query.scalars().all())
        
        rows = []
        STATE_CODES = {
            "01": "Jammu & Kashmir", "02": "Himachal Pradesh", "03": "Punjab", "04": "Chandigarh", "05": "Uttarakhand",
            "06": "Haryana", "07": "Delhi", "08": "Rajasthan", "09": "Uttar Pradesh", "10": "Bihar",
            "11": "Sikkim", "12": "Arunachal Pradesh", "13": "Nagaland", "14": "Manipur", "15": "Mizoram",
            "16": "Tripura", "17": "Meghalaya", "18": "Assam", "19": "West Bengal", "20": "Jharkhand",
            "21": "Odisha", "22": "Chhattisgarh", "23": "Madhya Pradesh", "24": "Gujarat", "26": "Dadra and Nagar Haveli and Daman and Diu",
            "27": "Maharashtra", "29": "Karnataka", "30": "Goa", "31": "Lakshadweep", "32": "Kerala",
            "33": "Tamil Nadu", "34": "Puducherry", "35": "Andaman & Nicobar Islands", "36": "Telangana", "37": "Andhra Pradesh",
            "38": "Ladakh"
        }

        for inv in invoices:
            # Gather branch company state details
            q_br = await db.execute(select(Branch).filter(Branch.id == inv.branch_id))
            branch = q_br.scalar_one_or_none()
            company_state = "22"
            if branch:
                q_comp = await db.execute(select(Company).filter(Company.id == branch.company_id))
                company = q_comp.scalar_one_or_none()
                if company:
                    company_state = company.state_code if company.state_code else (company.gstin[:2] if company.gstin else "22")
            
            # Aggregate payment details
            total_paid = sum([p.amount_paid for p in inv.payments])
            outstanding_amount = max(0.0, inv.total_amount - total_paid)
            
            payment_date = ""
            payment_mode = ""
            if inv.payments:
                sorted_payments = sorted(inv.payments, key=lambda x: x.payment_date, reverse=True)
                payment_date = sorted_payments[0].payment_date.strftime("%Y-%m-%d")
                payment_mode = sorted_payments[0].payment_mode.upper()
                
            # Customer details
            cust_name = "Unknown"
            cust_gstin = ""
            cust_address = ""
            customer = None
            if inv.sales_order and inv.sales_order.customer:
                customer = inv.sales_order.customer
            elif inv.delivery_challan and inv.delivery_challan.customer:
                customer = inv.delivery_challan.customer
                
            if customer:
                cust_name = customer.name
                cust_gstin = customer.gstin or ""
                cust_address = customer.billing_address or ""
                
            # Place of supply
            place_of_supply = "Other"
            cust_state_code = cust_gstin[:2] if cust_gstin else ""
            if cust_state_code in STATE_CODES:
                place_of_supply = f"{STATE_CODES[cust_state_code]} ({cust_state_code})"
            elif cust_address:
                place_of_supply = cust_address.split(",")[-1].strip()
                
            is_interstate = False
            if cust_state_code and company_state:
                is_interstate = (cust_state_code != company_state)
            
            # Iterate invoice items
            for item in inv.items:
                product_name = item.product.name if item.product else "Unknown Product"
                hsn_code = item.product.hsn_code if item.product else ""
                
                taxable_value = item.amount
                discount = item.discount_amount
                
                # GST Breakdown for this item
                cgst_pct = 0.0
                cgst_amt = 0.0
                sgst_pct = 0.0
                sgst_amt = 0.0
                igst_pct = 0.0
                igst_amt = 0.0
                
                if is_interstate:
                    igst_pct = item.tax_rate
                    igst_amt = item.tax_amount
                else:
                    cgst_pct = item.tax_rate / 2.0
                    cgst_amt = item.tax_amount / 2.0
                    sgst_pct = item.tax_rate / 2.0
                    sgst_amt = item.tax_amount / 2.0
                    
                rows.append({
                    "id": str(item.id),
                    "invoice_id": str(inv.id),
                    "invoice_number": inv.invoice_number,
                    "invoice_date": inv.date.strftime("%Y-%m-%d"),
                    "payment_date": payment_date,
                    "payment_mode": payment_mode,
                    "customer_name": cust_name,
                    "customer_gstin": cust_gstin,
                    "place_of_supply": place_of_supply,
                    "hsn_code": hsn_code,
                    "product_description": product_name,
                    "taxable_value": round(taxable_value, 2),
                    "discount": round(discount, 2),
                    "cgst_pct": round(cgst_pct, 2),
                    "cgst_amount": round(cgst_amt, 2),
                    "sgst_pct": round(sgst_pct, 2),
                    "sgst_amount": round(sgst_amt, 2),
                    "igst_pct": round(igst_pct, 2),
                    "igst_amount": round(igst_amt, 2),
                    "total_tax": round(item.tax_amount, 2),
                    "total_invoice_value": round(inv.total_amount, 2),
                    "tds_pct": 0.0,
                    "tds_amount": 0.0,
                    "outstanding_amount": round(outstanding_amount, 2),
                    "remarks": inv.status
                })
                
        return rows

    @staticmethod
    def generate_sales_summary_excel(rows: List[Dict[str, Any]]):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from io import BytesIO

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sales Summary"

        ws.views.sheetView[0].showGridLines = True

        font_family = "Segoe UI"
        
        header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1B4332", end_color="1B4332", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        data_font = Font(name=font_family, size=10)
        zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        totals_font = Font(name=font_family, size=11, bold=True)
        totals_fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")

        thin_border_side = Side(border_style="thin", color="CBD5E1")
        thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
        
        double_bottom_border = Border(
            left=thin_border_side, right=thin_border_side, 
            top=thin_border_side, 
            bottom=Side(border_style="double", color="1B4332")
        )

        headers = [
            "Invoice Number", "Invoice Date", "Payment Date", "Payment Mode", 
            "Customer Name", "Customer GSTIN", "Place of Supply", "HSN/SAC Code", 
            "Product/Service Description", "Taxable Value", "Discount", "CGST %", 
            "CGST Amount", "SGST %", "SGST Amount", "IGST %", "IGST Amount", 
            "Total Tax", "Total Invoice Value", "TDS %", "TDS Amount", 
            "Outstanding Amount", "Remarks"
        ]

        ws.append(headers)
        ws.row_dimensions[1].height = 28
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        align_left = Alignment(horizontal="left", vertical="center")
        align_center = Alignment(horizontal="center", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")

        col_alignments = {
            1: align_left,
            2: align_center,
            3: align_center,
            4: align_center,
            5: align_left,
            6: align_center,
            7: align_left,
            8: align_center,
            9: align_left,
            10: align_right,
            11: align_right,
            12: align_right,
            13: align_right,
            14: align_right,
            15: align_right,
            16: align_right,
            17: align_right,
            18: align_right,
            19: align_right,
            20: align_right,
            21: align_right,
            22: align_right,
            23: align_left
        }

        currency_format = '₹#,##,##0.00'
        percent_format = '0.00"%"'

        row_num = 2
        for data in rows:
            row_data = [
                data["invoice_number"],
                data["invoice_date"],
                data["payment_date"] if data["payment_date"] else "-",
                data["payment_mode"] if data["payment_mode"] else "-",
                data["customer_name"],
                data["customer_gstin"] if data["customer_gstin"] else "-",
                data["place_of_supply"],
                data["hsn_code"] if data["hsn_code"] else "-",
                data["product_description"],
                data["taxable_value"],
                data["discount"],
                data["cgst_pct"] / 100.0,
                data["cgst_amount"],
                data["sgst_pct"] / 100.0,
                data["sgst_amount"],
                data["igst_pct"] / 100.0,
                data["igst_amount"],
                data["total_tax"],
                data["total_invoice_value"],
                data["tds_pct"] / 100.0,
                data["tds_amount"],
                data["outstanding_amount"],
                data["remarks"]
            ]
            
            ws.append(row_data)
            ws.row_dimensions[row_num].height = 20
            
            is_zebra = (row_num % 2 == 1)
            row_fill = zebra_fill if is_zebra else white_fill
            
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_num, column=col_idx)
                cell.font = data_font
                cell.fill = row_fill
                cell.alignment = col_alignments.get(col_idx, align_left)
                cell.border = thin_border
                
                if col_idx in [10, 11, 13, 15, 17, 18, 19, 21, 22]:
                    cell.number_format = currency_format
                elif col_idx in [12, 14, 16, 20]:
                    cell.number_format = percent_format
            
            row_num += 1

        ws.row_dimensions[row_num].height = 24
        
        total_label_cell = ws.cell(row=row_num, column=1, value="TOTAL")
        total_label_cell.font = totals_font
        total_label_cell.fill = totals_fill
        total_label_cell.alignment = align_center
        total_label_cell.border = double_bottom_border
        
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=9)
        for col_idx in range(2, 10):
            ws.cell(row=row_num, column=col_idx).fill = totals_fill
            ws.cell(row=row_num, column=col_idx).border = double_bottom_border

        for col_idx in range(10, 24):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.font = totals_font
            cell.fill = totals_fill
            cell.border = double_bottom_border
            
            col_letter = get_column_letter(col_idx)
            
            if col_idx in [10, 11, 13, 15, 17, 18, 19, 21, 22]:
                cell.value = f"=SUM({col_letter}2:{col_letter}{row_num - 1})"
                cell.number_format = currency_format
                cell.alignment = align_right
            elif col_idx in [12, 14, 16, 20]:
                cell.value = ""
                cell.alignment = align_center
            else:
                cell.value = ""
                cell.alignment = align_left

        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val = str(cell.value or '')
                if val.startswith("="):
                    val = "₹99,99,999.00"
                if len(val) > max_len:
                    max_len = len(val)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
